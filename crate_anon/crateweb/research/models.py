#!/usr/bin/env python

"""
crate_anon/crateweb/research/models.py

===============================================================================

    Copyright (C) 2015-2019 Rudolf Cardinal (rudolf@pobox.com).

    This file is part of CRATE.

    CRATE is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    CRATE is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with CRATE. If not, see <http://www.gnu.org/licenses/>.

===============================================================================

"""

from collections import OrderedDict
import datetime
import io
import logging
from typing import Any, Dict, Generator, Iterable, List, Optional, Tuple, Union
import weakref
import zipfile
import json
import pytz

from cardinal_pythonlib.dbfunc import dictfetchall, get_fieldnames_from_cursor
from cardinal_pythonlib.django.fields.jsonclassfield import JsonClassField
from cardinal_pythonlib.excel import excel_to_bytes
from cardinal_pythonlib.exceptions import add_info_to_exception
from cardinal_pythonlib.hash import (
    get_longest_supported_hasher_output_length,
    hash64,
)
from cardinal_pythonlib.json.serialize import (
    json_encode,
    METHOD_STRIP_UNDERSCORE,
    register_for_json,
)
from cardinal_pythonlib.reprfunc import simple_repr
from cardinal_pythonlib.sql.sql_grammar import format_sql, SqlGrammar
from cardinal_pythonlib.tsv import make_tsv_row
from cardinal_pythonlib.django.function_cache import django_cache_function
from django.db import connections, DatabaseError, models
from django.db.models import QuerySet
from django.conf import settings
from django.http.request import HttpRequest
from django.db.backends.utils import CursorWrapper
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from crate_anon.anonymise.models import PatientInfoConstants
from crate_anon.common.sql import (
    ColumnId,
    columns_to_table_column_hierarchy,
    escape_percent_for_python_dbapi,
    make_grammar,
    sql_string_literal,
    SqlArgsTupleType,
    TableId,
    translate_sql_qmark_to_percent,
    WhereCondition,
)
from crate_anon.crateweb.research.html_functions import (
    HtmlElementCounter,
    N_CSS_HIGHLIGHT_CLASSES,
    prettify_sql_html,
)
from crate_anon.crateweb.research.research_db_info import (
    RESEARCH_DB_CONNECTION_NAME,
    research_database_info,
    SingleResearchDatabase,
)
from crate_anon.crateweb.research.sql_writer import (
    add_to_select,
    SelectElement,
)
from crate_anon.crateweb.config.constants import ResearchDbInfoKeys

log = logging.getLogger(__name__)


# =============================================================================
# Hacking django-pyodbc-azure, to stop it calling cursor.nextset() every time
# you ask it to do cursor.fetchone()
# =============================================================================
# BUG in django-pyodbc-azure==1.10.4.0 (providing
# sql_server/*), 2017-02-17: this causes
# ProgrammingError "No results. Previous SQL was not a query."
# The problem relates to sql_server/pyodbc/base.py
# CursorWrapper.fetchone() calling self.cursor.nextset(); if
# you comment this out, it works fine.
# Related:
# - https://github.com/pymssql/pymssql/issues/98

DJANGO_PYODBC_AZURE_ENGINE = 'sql_server.pyodbc'


def replacement_sqlserver_pyodbc_cursorwrapper_fetchone(self) -> List[Any]:
    """
    A function to replace ``CursorWrapper.fetchone()`` in
    ``sql_server/pyodbc/base.py`` from ``django-pyodbc-azure``.
    This replacement function does not call ``cursor.nextset()``.
    """
    # log.debug("Using monkeypatched fetchone(); self: {}; self.cursor: "
    #           "{}".format(repr(self), repr(self.cursor)))
    row = self.cursor.fetchone()
    if row is not None:
        row = self.format_row(row)
    # BUT DO NOT CALL self.cursor.nextset()
    return row


def hack_django_pyodbc_azure_cursorwrapper() -> None:
    """
    Monkey-patch part of the ``sql_server.pyodbc`` library from
    ``django-pyodbc-azure``. It replaces the ``fetchone()`` method with a
    version that doesn't call ``cursor.nextset()`` automatically.

    **It looks like this becomes unnecessary in django-pyodbc-azure==2.0.6.1
    or similar, because the call to ``cursor.nextset()`` is now only performed
    ``if not self.connection.supports_mars``.**

    *Notes*

    - I thought I wanted to modify an *instance*, not a *class*
      (https://tryolabs.com/blog/2013/07/05/run-time-method-patching-python/).

    - To modify a class, we do ``SomeClass.method = newmethod``.

    - But to modify an instance, we use ``instance.method =
      types.MethodType(newmethod, instance)``.

    - However, it turned out the instance was actually part of a long chain
      of cursor wrappers, including the Django debug toolbar. Classes included
      ``debug_toolbar.panels.sql.tracking.NormalCursorWrapper``;
      ``django.db.backends.utils.CursorDebugWrapper``.
      And in any case, modifying the class is a sensible thing.

    """
    try:
        # noinspection PyUnresolvedReferences
        from sql_server.pyodbc.base import CursorWrapper
        log.info("Monkey-patching sql_server.pyodbc.base.CursorWrapper."
                 "fetchone to disable automatic call to cursor.nextset()")
        CursorWrapper.fetchone = replacement_sqlserver_pyodbc_cursorwrapper_fetchone  # noqa
    except ImportError:
        return


if getattr(settings, 'DISABLE_DJANGO_PYODBC_AZURE_CURSOR_FETCHONE_NEXTSET',
           True):
    # http://stackoverflow.com/questions/5601590/how-to-define-a-default-value-for-a-custom-django-setting  # noqa
    hack_django_pyodbc_azure_cursorwrapper()


@django_cache_function(timeout=None)
def database_last_updated(dbname: str) -> Optional[datetime.datetime]:
    """
    Returns a datetime object specifying when the database was last run.

    If there are any tables with a column specifying last updated time
    (based on the column name provided in the config file) but all dates
    are null, the function will return the minimum date possible. If there
    are no such tables, the function will return None.
    """
    try:
        dbinfo = research_database_info.get_dbinfo_by_name(dbname)
    except ValueError:
        raise ValueError(
            f"Database {dbname} is not specified in config file")
    tables_with_timecol = []
    for col in dbinfo.colinfolist:
        if col.column_name == dbinfo.update_date_field:
            tables_with_timecol.append(col.table_name)
    if not tables_with_timecol:
        return None
    latest_time = mindate = datetime.datetime.min
    for table in tables_with_timecol:
        # Not quite sure about the different dialects ...
        if dbinfo.rdb_info.dialect == 'mysql':
            sql = (
                f"SELECT {dbinfo.update_date_field} "
                f"FROM {dbinfo.schema_name}.{table}"
            )
        else:
            # Dialect must be mssql because ResearchDatabaseInfo checks if
            # supported dialect
            sql = (
                f"SELECT {dbinfo.update_date_field} "
                f"FROM {dbinfo.database}.{dbinfo.schema_name}.{table}"
            )
        with get_executed_researchdb_cursor(sql) as cursor:
            times = cursor.fetchall()
            times = [t[0] if t[0] else mindate for t in times]
            times.append(latest_time)
            latest_time = max(times)
    return latest_time


@django_cache_function(timeout=None)
def last_updated_all_dbs() -> datetime.datetime:
    """
    Returns a datetime object specifying the most recent databse update, or
    None if there is no information on this.
    """
    last_updated_all = datetime.datetime.min
    for db in settings.RESEARCH_DB_INFO:
        db_last_updated = database_last_updated(db[ResearchDbInfoKeys.NAME])
        if db_last_updated is None:
            continue
        last_updated_all = max(last_updated_all, db_last_updated)
    if last_updated_all == datetime.datetime.min:
        # This will only happen if none of the databases have update time
        # fields
        last_updated_all = None
    return last_updated_all


# =============================================================================
# Cursors
# =============================================================================

def debug_query() -> None:
    """
    Executes a test query that just selects a constant, using the research
    database (i.e. ``connections['research']``).
    """
    cursor = connections[RESEARCH_DB_CONNECTION_NAME].cursor()
    cursor.execute("SELECT 'debug'")


def get_executed_researchdb_cursor(sql: str,
                                   args: List[Any] = None) -> CursorWrapper:
    """
    Executes a query on the research database. Returns a wrapped cursor that
    can be used as a context manager that will close the cursor on completion.

    Args:
        sql: SQL text
        args: arguments to SQL query

    Returns:
        a :class:`django.db.backends.utils.CursorWrapper`, which is a context
        manager that behaves as the executed cursor and also closes it on
        completion
        
    Test code:
    
    .. code-block:: python

        import os
        import django
        os.environ['DJANGO_SETTINGS_MODULE'] = 'crate_anon.crateweb.config.settings'
        django.setup()
        from crate_anon.crateweb.research.models import *
        c = get_executed_researchdb_cursor("SELECT 1")

    """  # noqa
    args = args or []
    cursor = connections[RESEARCH_DB_CONNECTION_NAME].cursor()  # type: CursorWrapper  # noqa
    try:
        cursor.execute(sql, args or None)
    except DatabaseError as exception:
        add_info_to_exception(exception, {'sql': sql, 'args': args})
        raise
    return cursor


# =============================================================================
# Data going to Excel files
# =============================================================================

ILLEGAL_CHARACTERS_REPLACED_WITH = ""


def gen_excel_row_elements(worksheet: Worksheet,
                           row: Iterable) -> Generator[Any, None, None]:
    r"""
    Given an Excel worksheet row, generate individual cell contents, cell by
    cell.

    Args:
        worksheet: a :class:`openpyxl.worksheet.worksheet.Worksheet`, which we
            need in order to find the worksheet's encoding
        row: the row to iterate through

    Yields:
        the contents of each cell

    Reasons for this function:

    1. We need a tuple/list/generator, as ``openpyxl`` checks its types
       manually.

      - We want to have a Worksheet object from ``openpyxl``, and say something
        like

        .. code-block:: python

            ws.append(row)

        where "row" has come from a database query.

      - However, ``openpyxl`` doesn't believe in duck-typing; see
        ``Worksheet.append()`` in ``openpyxl/worksheet/worksheet.py``. So
        sometimes the plain append works (e.g. from MySQL results), but
        sometimes it fails, e.g. when the row is of type ``pyodbc.Row``.

      - So we must coerce it to a tuple, list, or generator.

      - A generator will be the most efficient.

    2. If a string fails certain checks, openpyxl will raise an
       :exc:`IllegalCharacterError` exception. We need to work around that.
       We'll use the "forgiveness, not permission" maxim. Specifically, it
       dislikes strings matching its ``ILLEGAL_CHARACTERS_RE``, which contains
       unprintable low characters matching this:

        .. code-block:: python

            r'[\000-\010]|[\013-\014]|[\016-\037]'

        Note the use of octal; ``\037`` is decimal 31.

        ``openpyxl`` gets to its ``Cell.check_string()`` function for these
        types:

        .. code-block:: python

            STRING_TYPES = (basestring, unicode, bytes)

        In Python 3, this means ``(str, str, bytes)``. So we should check
        ``str`` and ``bytes``. (For ``bytes``, we'll follow its method of
        converting to ``str`` in the encoding of the worksheet's choice.)
    """
    # Docstring must be a raw string for Sphinx! See
    # http://openalea.gforge.inria.fr/doc/openalea/doc/_build/html/source/sphinx/rest_syntax.html#text-syntax-bold-italic-verbatim-and-special-characters  # noqa
    for element in row:
        if isinstance(element, bytes):
            # Convert to str using the worksheet's encoding.
            element = element.decode(worksheet.encoding)
            # ... or: str(element, encoding)

        if isinstance(element, str):
            yield ILLEGAL_CHARACTERS_RE.sub(ILLEGAL_CHARACTERS_REPLACED_WITH,
                                            element)
        else:
            yield element


# =============================================================================
# Query highlighting class
# =============================================================================

HIGHLIGHT_FWD_REF = "Highlight"


class Highlight(models.Model):
    """
    Represents the highlighting of a query.
    """
    id = models.AutoField(primary_key=True)  # automatic
    user = models.ForeignKey(settings.AUTH_USER_MODEL,
                             on_delete=models.CASCADE)
    colour = models.PositiveSmallIntegerField(verbose_name="Colour number")
    text = models.CharField(max_length=255, verbose_name="Text to highlight")
    active = models.BooleanField(default=True)

    def __str__(self) -> str:
        return f"colour={self.colour}, text={self.text}"

    def get_safe_colour(self) -> int:
        """
        Returns our ``colour`` attribute, coerced to the range ``[0,
        N_CSS_HIGHLIGHT_CLASSES - 1]`` (inclusive).
        """
        if self.colour is None:
            return 0
        return max(0, min(self.colour, N_CSS_HIGHLIGHT_CLASSES - 1))

    @staticmethod
    def as_ordered_dict(highlight_list: Iterable[HIGHLIGHT_FWD_REF]) \
            -> Dict[int, List[HIGHLIGHT_FWD_REF]]:
        """
        Converts a iterable of :class:`Highlight` objects into a dictionary
        that collects them by highlight number.

        Args:
            highlight_list: list of :class:`Highlight` objects

        Returns:
            an OrderedDict whose keys are highlight colour numbers (in
            ascending order), and whose values are lists of all the
            :class:`Highlight` objects using that highlight colour number

        """
        d = dict()  # type: Dict[int, List[HIGHLIGHT_FWD_REF]]
        for highlight in highlight_list:
            n = highlight.get_safe_colour()
            if n not in d:
                d[n] = []  # type: List[HIGHLIGHT_FWD_REF]
            d[n].append(highlight)
        # noinspection PyTypeChecker
        return OrderedDict(sorted(d.items()))

    @staticmethod
    def get_active_highlights(request: HttpRequest) -> QuerySet:
        """
        Return all active highlights for the current user.

        Args:
            request: the :class:`django.http.request.HttpRequest`

        Returns:
            a :class:`django.db.models.QuerySet` of the :class:`Highlight`
            objects

        """
        return Highlight.objects.filter(user=request.user, active=True)

    def activate(self) -> None:
        """
        Mark this highlight as active.
        """
        self.active = True
        self.save()

    def deactivate(self) -> None:
        """
        Mark this highlight as inactive.
        """
        self.active = False
        self.save()


# =============================================================================
# Query classes
# =============================================================================

QUERY_FWD_REF = "Query"


class QueryBase(models.Model):
    """
    Abstract base class for the two query classes.
    """
    class Meta:
        abstract = True
        app_label = "research"

    id = models.AutoField(primary_key=True)  # automatic

    sql = models.TextField(verbose_name='SQL query')
    sql_hash = models.BigIntegerField(
        verbose_name='64-bit non-cryptographic hash of SQL query')
    args = JsonClassField(verbose_name='SQL arguments (as JSON)', null=True)
    # ... https://github.com/shrubberysoft/django-picklefield
    raw = models.BooleanField(
        default=False, verbose_name='SQL is raw, not parameter-substituted')
    qmark = models.BooleanField(
        default=True,
        verbose_name='Parameter-substituted SQL uses ?, not %s, '
        'as placeholders')
    # active = models.BooleanField(default=True)  # see save() below
    created = models.DateTimeField(auto_now_add=True)
    deleted = models.BooleanField(
        default=False,
        verbose_name="Deleted from the user's perspective. "
                     "Audited queries are never properly deleted.")

    def __repr__(self) -> str:
        return simple_repr(self, ['id', 'sql', 'args', 'raw', 'qmark',
                                  'created', 'deleted'])

    # -------------------------------------------------------------------------
    # SQL queries
    # -------------------------------------------------------------------------

    def get_original_sql(self) -> str:
        """
        Returns the stored raw SQL.
        """
        # noinspection PyTypeChecker
        return self.sql


def _close_cursor(cursor: Optional[CursorWrapper]) -> None:
    if cursor:
        # log.debug("Closing cursor")
        cursor.close()


class Query(QueryBase):
    """
    Class to query the research database.
    """
    NO_NULL = "_no_null"  # special output

    user = models.ForeignKey(settings.AUTH_USER_MODEL,
                             on_delete=models.CASCADE)
    active = models.BooleanField(default=True)  # see save() below
    audited = models.BooleanField(default=False)
    display = models.TextField(
        default="[]",
        verbose_name="Subset of output columns to be displayed")
    no_null = models.BooleanField(
        default=False,
        verbose_name="Omit Null columns for this query when displayed")
    last_run = models.DateTimeField(null=True, default=None)

    def __init__(self, *args, **kwargs) -> None:
        """
        Initialize our cache.
        """
        super().__init__(*args, **kwargs)
        self._executed_cursor = None  # type: CursorWrapper
        self._column_names = None  # type: List[str]
        self._rowcount = None  # type: int
        self._rows = None  # type: List[List[Any]]
        self._display_list = None  # type: List[str]
        self._display_indexes = None  # type: List[int]
        self._n_times_executed = 0
        self._finalizer = None

    def activate(self) -> None:
        """
        Activate this query (and deactivates any others).
        """
        self.active = True
        self.save()

    def __repr__(self) -> str:
        return simple_repr(self, ['id', 'user', 'sql', 'args', 'raw', 'qmark',
                                  'active', 'created', 'deleted', 'audited'])

    def save(self, *args, **kwargs) -> None:
        """
        Custom save method. Ensures that only one :class:`Query` has ``active
        == True`` for a given user. Also sets the hash.
        """
        # http://stackoverflow.com/questions/1455126/unique-booleanfield-value-in-django  # noqa
        if self.active:
            Query.objects.filter(user=self.user, active=True)\
                         .update(active=False)
        self.sql_hash = hash64(self.sql)
        super().save(*args, **kwargs)

    # -------------------------------------------------------------------------
    # SQL queries
    # -------------------------------------------------------------------------

    def get_sql_args_for_django(self) -> Tuple[str, Optional[List[Any]]]:
        """
        Get sql/args in a format suitable for Django, with ``%s`` placeholders,
        or as escaped raw SQL.

        Returns:
            tuple: ``sql, args``

        - If :attr:`raw` is set, return our raw SQL with ``%`` escaped to
          ``%%``;
        - otherwise, if :attr:`qmark` is set, return our raw SQL with ``?``
          argument placeholders translated to ``%s`` argument placeholders;
        - otherwise, return the raw SQL.

        """
        if self.raw:
            # noinspection PyTypeChecker
            sql = escape_percent_for_python_dbapi(self.sql)
            args = None
        else:
            if self.qmark:
                # noinspection PyTypeChecker
                sql = translate_sql_qmark_to_percent(self.sql)
            else:
                sql = self.sql
            args = self.args
        return sql, args

    def get_executed_cursor(self) -> CursorWrapper:
        """
        Get cursor with a query executed (based on our attributes :attr:`sql`,
        :attr:`args`, :attr:`raw`, :attr:`qmark`).

        Returns:
            a :class:`django.db.backends.utils.CursorWrapper`

        Do NOT use this with ``with``, as in:

        .. code-block:: python

            with query.get_executed_cursor() as cursor:
                # do stuff

        You could do that (and in general it's what Django advises) but we are
        trying to be fancy here and use the cursor more efficiently.

        """
        if self._executed_cursor is None:
            sql, args = self.get_sql_args_for_django()
            cursor = get_executed_researchdb_cursor(sql, args)
            self._n_times_executed += 1
            # log.debug("Query: n_times_executed: {}".format(
            #     self._n_times_executed))
            # log.debug("\n" + "".join(traceback.format_stack()))
            if self._n_times_executed > 1:
                log.warning(f"Inefficient: Query executed "
                            f"{self._n_times_executed} times")
            try:
                # noinspection PyTypeChecker
                self._column_names = get_fieldnames_from_cursor(cursor)
            except TypeError:
                self._column_names = []
            self._rowcount = cursor.rowcount
            self._executed_cursor = cursor
            self._finalizer = weakref.finalize(
                self, _close_cursor, self._executed_cursor)
        return self._executed_cursor

    def _invalidate_executed_cursor(self) -> None:
        """
        Mark the executed cursor as dead (e.g. iterated through).
        """
        if self._executed_cursor is not None:
            self._finalizer()
            self._finalizer = None
        self._executed_cursor = None

    def _cache_basics(self) -> None:
        """
        Cache rowcount and column names.

        Raises:
            :exc:`DatabaseError` if the query fails
        """
        if self._rowcount is None:
            self.get_executed_cursor()  # will cache

    def _cache_all(self) -> None:
        """
        Fetch everything from the query and cache it.

        Raises:
            :exc:`DatabaseError` if the query fails
        """
        if self._rows is None:
            cursor = self.get_executed_cursor()
            self._rows = cursor.fetchall()
            self._invalidate_executed_cursor()

    def get_column_names(self) -> List[str]:
        """
        Returns column names from the query's cursor.

        Raises:
            :exc:`DatabaseError` if the query fails
        """
        if self._column_names is None:
            self._cache_basics()
        return self._column_names

    def get_rowcount(self) -> int:
        """
        Returns the rowcount from the cursor.

        Raises:
            :exc:`DatabaseError` if the query fails
        """
        if self._rowcount is None:
            self._cache_basics()
        return self._rowcount

    def get_rows(self) -> List[List[Any]]:
        """
        Returns all rows from the query, as a list.

        Raises:
            :exc:`DatabaseError` if the query fails
        """
        self._cache_all()
        return self._rows

    def gen_rows(self) -> Generator[List[Any], None, None]:
        """
        Generate rows from the query.

        Raises:
            :exc:`DatabaseError` if the query fails
        """
        if self._rows is None:
            # No cache
            cursor = self.get_executed_cursor()
            row = cursor.fetchone()
            while row is not None:
                yield row
                row = cursor.fetchone()
            self._invalidate_executed_cursor()
        else:
            # Cache
            for row in self._rows:
                yield row

    def dictfetchall(self) -> List[Dict[str, Any]]:
        """
        Executes the query and returns all results as a list of OrderedDicts
        (one for each row, mapping column names to values).

        Raises:
            :exc:`DatabaseError` if the query fails
        """
        if self._rows is None:
            # No cache
            cursor = self.get_executed_cursor()
            # noinspection PyTypeChecker
            result = dictfetchall(cursor)
            self._invalidate_executed_cursor()
            return result
        else:
            # Cache
            columns = self._column_names
            return [
                OrderedDict(zip(columns, row))
                for row in self._rows
            ]

    def update_last_run(self) -> None:
        self.last_run = datetime.datetime.now()
        self.save()

    @property
    def run_since_update(self) -> Optional[bool]:
        # Currently doesn't check which databases are involved in the
        # query - just checks all databases
        last_updated_all = last_updated_all_dbs()
        if last_updated_all is None:  # no info from dbs
            return None
        elif self.last_run is None:  # query never run
            return False
        else:
            # Make last_updated_all timezone aware so they can be compared
            last_updated_all = pytz.utc.localize(last_updated_all)
            return self.last_run > last_updated_all

    # -------------------------------------------------------------------------
    # Fetching
    # -------------------------------------------------------------------------

    @staticmethod
    def get_active_query_or_none(request: HttpRequest) \
            -> Optional[QUERY_FWD_REF]:
        """
        Returns the active query for this user, or ``None``.

        Args:
            request: the :class:`django.http.request.HttpRequest`

        Returns:
            a :class:`Query`, or ``None``.

        """
        if not request.user.is_authenticated:
            return None
        try:
            return Query.objects.get(user=request.user, active=True)
        except Query.DoesNotExist:
            return None

    @staticmethod
    def get_active_query_id_or_none(request: HttpRequest) -> Optional[int]:
        """
        Returns the active query's integer ID for this user, or ``None``.

        Args:
            request: the :class:`django.http.request.HttpRequest`

        Returns:
            the active query's integer PK, or ``None``.

        """
        if not request.user.is_authenticated:
            return None
        try:
            query = Query.objects.get(user=request.user, active=True)
            return query.id
        except Query.DoesNotExist:
            return None

    # -------------------------------------------------------------------------
    # Activating, deleting, auditing
    # -------------------------------------------------------------------------

    # This isn't needed in the base class because it only applies to
    # audited queries
    def mark_deleted(self) -> None:
        """
        Mark the query as deleted.

        This will stop it being shown. It will not delete it from the database.

        We use this deletion method for queries that have been executed, so
        need an audit trail.
        """
        if self.deleted:
            # log.debug("pointless")
            return
        self.deleted = True
        self.active = False
        # log.debug("about to save")
        self.save()
        # log.debug("saved")

    def mark_audited(self) -> None:
        """
        Mark the query as having been executed and audited. (This prevents it
        from being wholly deleted.)
        """
        if self.audited:
            return
        self.audited = True
        self.save()

    def audit(self, count_only: bool = False, n_records: int = 0,
              failed: bool = False, fail_msg: str = "") -> None:
        """
        Audit the execution of this query:

        - insert an audit entry referring to this query
        - mark the query as having been audited (so it's not deleted)

        Args:
            count_only: did we know (in advance) that this was a
                ``COUNT()``-only query?
            n_records: how many records were returned?
            failed: did the query fail?
            fail_msg: if the query failed, the associated failure message
        """
        a = QueryAudit(query=self,
                       count_only=count_only,
                       n_records=n_records,
                       failed=failed,
                       fail_msg=fail_msg)
        a.save()
        self.mark_audited()

    def delete_if_permitted(self) -> None:
        """
        Delete the query.

        - If a query has been executed and therefore audited, it isn't properly
          deleted; it's just marked as deleted.
        - If a query has never been executed, we can delete it entirely.
        """
        if self.deleted:
            log.debug("already flagged as deleted")
            return
        if self.audited:
            log.debug("marking as deleted")
            self.mark_deleted()
        else:
            # actually delete
            log.debug("actually deleting")
            self.delete()

    # -------------------------------------------------------------------------
    # Filtering columns for display output
    # -------------------------------------------------------------------------

    def set_display_list(self, display_list: List[str]) -> None:
        """
        Sets the internal JSON field, stored in the database, from a list of
        column headings to display.

        Args:
            display_list: list of columns to display
        """
        self.display = json.dumps(display_list)
        self._display_list = None  # clear cache

    def _get_display_list(self) -> List[str]:
        """
        Returns a list of columns to display, from our internal JSON
        representation.
        """
        if not self.display:
            return []
        try:
            result = json.loads(self.display)
        except json.decoder.JSONDecodeError:  # e.g. junk
            log.warning("Query.display field: bad JSON, returning []")
            return []
        # Now check it's a list of str:
        if not isinstance(result, list):
            log.warning("Query.display field: not a list, returning []")
            return []
        if not all(isinstance(x, str) for x in result):
            log.warning("Query.display field: contains non-strings, "
                        "returning []")
            return []
        return result

    def get_display_list(self) -> List[str]:
        """
        Returns a list of columns to display, from our internal JSON
        representation. Uses :func:`_get_display_list` and caches it.
        """
        if self._display_list is None:
            self._display_list = self._get_display_list()
        # log.debug("Query.get_display_list() -> {!r}".format(
        #     self._display_list))
        return self._display_list

    def _get_display_indexes(self) -> Optional[List[int]]:
        """
        Returns the indexes of the result columns that we wish to display.

        Raises:
            :exc:`DatabaseError` on query failure
        """
        display_fieldnames = self.get_display_list()
        # If the display attribute is empty assume the user wants all fields
        select_all = not display_fieldnames

        if self.no_null:
            self._cache_all()  # writes to self._rows

        all_column_names = self.get_column_names()

        if select_all and not self.no_null:
            # No filtering. Provide the original indexes quickly.
            return list(range(len(all_column_names)))

        field_indexes = []  # type: List[int]
        # Do this to make sure included fields are actually in the results
        for i, name in enumerate(all_column_names):
            if select_all or name in display_fieldnames:
                if self.no_null:
                    # Exclude fields where all values are null, if no_null
                    # is switched on.
                    for row in self._rows:
                        if row[i] is not None:
                            field_indexes.append(i)
                            break
                else:
                    field_indexes.append(i)
        return field_indexes

    def get_display_indexes(self) -> Optional[List[int]]:
        """
        Returns the indexes of the result columns that we wish to display.
        Uses :func:`_get_display_indexes` and caches it.

        Raises:
            :exc:`DatabaseError` on query failure
        """
        if self._display_indexes is None:
            self._display_indexes = self._get_display_indexes()
        return self._display_indexes

    def get_display_column_names(self) -> List[str]:
        """
        Returns the filtered column names.
        """
        column_names = self.get_column_names()
        display_indexes = self.get_display_indexes()
        return [column_names[i] for i in display_indexes]

    def gen_display_rows(self) -> Generator[List[Any], None, None]:
        """
        Generates all filtered rows.
        """
        field_indexes = self.get_display_indexes()
        if not field_indexes:
            # No columns specifically selected; return all columns
            for row in self.gen_rows():
                yield row
        else:
            for row in self.gen_rows():
                yield [row[i] for i in field_indexes]

    def get_display_rows(self) -> List[List[Any]]:
        """
        Returns a list of all filtered rows.
        """
        if self._rows is not None:
            # Pre-cached; there may be a shortcut
            field_indexes = self.get_display_indexes()
            if not field_indexes:
                # No columns specifically selected; return all columns
                return self._rows
        # Otherwise, use the generator:
        return list(self.gen_display_rows())

    def make_tsv(self) -> str:
        """
        Executes the query and returns a TSV result (as a multiline string).
        """
        fieldnames = self.get_display_column_names()
        tsv = make_tsv_row(fieldnames)
        for row in self.gen_display_rows():
            tsv += make_tsv_row(row)
        self.update_last_run()
        return tsv

    def make_excel(self) -> bytes:
        """
        Executes the query and returns an Excel workbook, in binary.
        """
        self._cache_all()
        wb = Workbook()
        wb.remove_sheet(wb.active)  # remove the autocreated blank sheet
        sheetname = f"query_{self.id}"
        ws = wb.create_sheet(sheetname)
        now = datetime.datetime.now()

        fieldnames = self.get_display_column_names()
        ws.append(fieldnames)
        for row in self.gen_display_rows():
            ws.append(gen_excel_row_elements(ws, row))

        sql_ws = wb.create_sheet(title="SQL")
        sql_ws.append(["SQL", "Executed_at"])
        sql_ws.append([self.get_original_sql(), now])
        self.update_last_run()
        return excel_to_bytes(wb)


class SitewideQuery(QueryBase):
    """
    Class representing a site-wide query for research database.

    - Site-wide queries are not attached to any particular user.
    - They are templatized with placeholders.
    - Placeholders begin with ``[[`` and end with ``]]``.
    - The user is asked to fill in values for the placeholders.

    """

    description = models.TextField(verbose_name='query description',
                                   default="")

    @property
    def sql_chunks(self) -> List[str]:
        """
        Returns a list of SQL chunks and placeholders made from the original
        SQL. Placeholders begin with ``[[`` and end with ``]]``.

        For example, if the sql is

        .. code-block:: none

            SELECT * FROM [[table]] WHERE brcid="[[brcid]]";

        then ``sql_chunks`` will be

        .. code-block:: python

            [
                'SELECT * FROM ',
                'table',
                ' WHERE brcid="',
                'brcid',
                '";'
            ]

        Note that the first element (and all elements with even [zero-based]
        list indexes) are SQL, not placeholders. All elements with odd indexes
        are placeholders.
        """
        sql_string = self.sql
        placeholder_start = "[["
        placeholder_end = "]]"
        startlen = len(placeholder_start)
        endlen = len(placeholder_end)
        chunks = []  # type: List[str]
        index1 = sql_string.find(placeholder_start)
        index2 = sql_string.find(placeholder_end)
        while index1 != -1 and index2 != -1:  # placeholder present
            # get bit of sql up to next '[['
            chunk = sql_string[:index1]
            # get bit of sql between '[[' and ']]'
            placeholder = sql_string[index1 + startlen:index2]
            chunks.append(chunk)
            chunks.append(placeholder)
            # get bit of sql after '[[' - this forms new substring to check
            sql_string = sql_string[index2 + endlen:]
            index1 = sql_string.find(placeholder_start)
            index2 = sql_string.find(placeholder_end)
        # Deal with any remainder
        if sql_string:
            chunks.append(sql_string)

        return chunks

    def save(self, *args, **kwargs) -> None:
        """
        Custom save method. Sets the hash.
        """
        self.sql_hash = hash64(self.sql)
        super().save(*args, **kwargs)


# =============================================================================
# Query auditing class
# =============================================================================

class QueryAudit(models.Model):
    """
    Audit log for a query.
    """
    id = models.AutoField(primary_key=True)  # automatic
    query = models.ForeignKey('Query', on_delete=models.PROTECT)
    when = models.DateTimeField(auto_now_add=True)
    count_only = models.BooleanField(default=False)
    n_records = models.IntegerField(default=0)
    # ... not PositiveIntegerField; SQL Server gives -1, for example
    failed = models.BooleanField(default=False)
    fail_msg = models.TextField()

    def __str__(self):
        return f"<QueryAudit id={self.id}>"


# =============================================================================
# Lookup class for secret RID-to-PID conversion
# =============================================================================

# class PidLookupRouter(object):
#     # https://docs.djangoproject.com/en/1.8/topics/db/multi-db/
#     # https://newcircle.com/s/post/1242/django_multiple_database_support
#     # noinspection PyMethodMayBeStatic,PyUnusedLocal
#     def db_for_read(self, model: Type[models.Model], **hints) -> Optional[str]:  # noqa
#         """
#         read model PidLookup -> look at database secret
#         """
#         # log.debug("PidLookupRouter: {}".format(model._meta.model_name))
#         # if model._meta.model_name == PidLookup._meta.model_name:
#         if model == PidLookup:
#             return 'secret'
#         return None
#
#     # noinspection PyUnusedLocal
#     @staticmethod
#     def allow_migrate(db: str, app_label: str, model_name: str = None,
#                       **hints) -> bool:
#         # 2017-02-12, to address bug:
#         # - https://code.djangoproject.com/ticket/27054
#         # See also:
#         # - https://docs.djangoproject.com/en/1.10/topics/db/multi-db/#using-other-management-commands  # noqa
#         return db == 'default'


class PidLookup(models.Model):
    """
    Lookup class for secret RID-to-PID conversion.

    - Used via one or other of the 'secret' database connections.
    - Intended for READ-ONLY access to that table.

    - Since we have fixed the tablenames for the anonymiser, we remove the
      ``settings.SECRET_MAP`` option. See
      :class:`crate_anon.anonymise.models.PatientInfo`. Moreover, we fix the
      maximum length, regardless of the specifics of the config used.

    - Use as e.g. ``Lookup(pid=XXX)``.

    """
    pid = models.PositiveIntegerField(
        primary_key=True,
        db_column=PatientInfoConstants.PID_FIELDNAME)
    mpid = models.PositiveIntegerField(
        db_column=PatientInfoConstants.MPID_FIELDNAME)
    rid = models.CharField(
        db_column=PatientInfoConstants.RID_FIELDNAME,
        max_length=get_longest_supported_hasher_output_length())
    mrid = models.CharField(
        db_column=PatientInfoConstants.MRID_FIELDNAME,
        max_length=get_longest_supported_hasher_output_length())
    trid = models.PositiveIntegerField(
        db_column=PatientInfoConstants.TRID_FIELDNAME)

    class Meta:
        managed = False
        db_table = PatientInfoConstants.SECRET_MAP_TABLENAME

    # https://stackoverflow.com/questions/12158463/how-can-i-make-a-model-read-only  # noqa
    def save(self, *args, **kwargs) -> None:
        return

    def delete(self, *args, **kwargs) -> None:
        return


def get_pid_lookup(dbinfo: SingleResearchDatabase,
                   pid: Union[int, str] = None,
                   mpid: Union[int, str] = None,
                   trid: int = None,
                   rid: str = None,
                   mrid: str = None) -> Optional[PidLookup]:
    """
    Looks up a patient in the secret lookup database associated with a
    database, from one of several possible identifiers.
    
    Args:
        dbinfo: a
            :class:`crate_anon.crateweb.research.research_db_info.SingleResearchDatabase`
        pid: optional patient identifier (PID) value
        mpid: optional master patient identifier (MPID) value
        trid: optional transient research identifier (TRID) value
        rid: optional research identifier (RID) value
        mrid: optional master research identifier (MRID) value

    Returns:
        a :class:`crate_anon.crateweb.research.models.PidLookup` or ``None``
        
    Raises:
        :exc:`ValueError` if none of the IDs was specified

    """  # noqa
    dbalias = dbinfo.secret_lookup_db
    assert dbalias
    q = PidLookup.objects.using(dbalias)
    if trid is not None:
        lookup = q.get(trid=trid)
    elif rid is not None:
        lookup = q.get(rid=rid)
    elif mrid is not None:
        lookup = q.get(mrid=mrid)
    elif pid is not None:
        lookup = q.get(pid=pid)
    elif mpid is not None:
        lookup = q.get(mpid=mpid)
    else:
        raise ValueError("no input")
    return lookup


def get_mpid(dbinfo: SingleResearchDatabase,
             trid: int = None,
             rid: str = None,
             mrid: str = None) -> int:
    """
    Returns the MPID for a patient, looked up from one of the research IDs.

    Args:
        dbinfo: a
            :class:`crate_anon.crateweb.research.research_db_info.SingleResearchDatabase`
        trid: optional transient research identifier (TRID) value
        rid: optional research identifier (RID) value
        mrid: optional master research identifier (MRID) value

    Returns:
        the integer MPID, or ``None``

    Raises:
        :exc:`ValueError` if none of the IDs was specified
    """  # noqa
    lookup = get_pid_lookup(dbinfo=dbinfo, trid=trid, rid=rid, mrid=mrid)
    # noinspection PyTypeChecker
    return lookup.mpid


def get_pid(dbinfo: SingleResearchDatabase,
            trid: int = None,
            rid: str = None,
            mrid: str = None) -> int:
    """
    Returns the PID for a patient, looked up from one of the research IDs.

    Args:
        dbinfo: a
            :class:`crate_anon.crateweb.research.research_db_info.SingleResearchDatabase`
        trid: optional transient research identifier (TRID) value
        rid: optional research identifier (RID) value
        mrid: optional master research identifier (MRID) value

    Returns:
        the integer PID, or ``None``

    Raises:
        :exc:`ValueError` if none of the IDs was specified
    """  # noqa
    lookup = get_pid_lookup(dbinfo=dbinfo, trid=trid, rid=rid, mrid=mrid)
    # noinspection PyTypeChecker
    return lookup.pid


# =============================================================================
# Patient Explorer multi-query classes
# =============================================================================

class TableQueryArgs(object):
    """
    Represents SQL for a specific table, with arguments for the SQL. Used by
    :class:`PatientMultiQuery`.
    """
    def __init__(self,
                 table_id: TableId,
                 sql: str,
                 args: List[Any]) -> None:
        """
        Args:
            table_id: a :class:`crate_anon.common.sql.TableId` that this query
                is selecting from
            sql: SQL text
            args: a list of arguments to the SQL
        """
        self.table_id = table_id
        self.sql = sql
        self.args = args


@register_for_json(method=METHOD_STRIP_UNDERSCORE)
class PatientMultiQuery(object):
    """
    Represents a set of queries across many tables relating to one or several
    patients (but the same patients across all the tables).
    
    Used for the Patient Explorer.
    
    *Development notes:*
        
    - Patient ID query
    
      - Single database is easy; we can use RID or TRID, and therefore TRID for
        performance.
        
        Note that ``UNION`` gives only ``DISTINCT`` results by default (``UNION
        ALL`` gives everything); see
        http://stackoverflow.com/questions/49925/what-is-the-difference-between-union-and-union-all.
        
        .. code-block:: sql
    
            -- Clear, but extensibility of boolean logic less clear:
            SELECT trid
                FROM diagnosis_table
                WHERE diagnosis LIKE 'F20%'
            INTERSECT
            SELECT trid
                FROM progress_note_table
                WHERE note LIKE '%schizophreni%' OR note LIKE '%depression%'
            ORDER BY trid
            -- ... logic across tables requires careful arrangement of UNION vs. INTERSECT
            -- ... logic for multiple fields within one table can be done with AND/OR
    
            -- Slower (?), but simpler to manipulate logic?
            SELECT DISTINCT something.trid
            FROM diagnosis_table INNER JOIN progress_note_table
            ON diagnosis_table.trid = progress_note_table.trid
            WHERE
                diagnosis_table.diagnosis LIKE 'F20%'
                AND (progress_note_table.note LIKE '%schizophreni%'
                     OR progress_note_table.notenote LIKE '%depression%')
            ORDER BY something.trid
            -- ... boolean logic can all be encapsulated in a single WHERE clause
            -- ... can also share existing join code
            -- ... ?reasonable speed since the TRID fields will be indexed
            -- ... preferable.
    
    - Which ID for the patient ID query?
    
      - the TRID (for speed, inc. sorting) of the first database
      - can use the TRID from the first "where clause" table
        (don't have to join to a master patient table)
      - join everything across databases as before
    
    - Results queries
    
      .. code-block:: none
    
        -- Something like:
    
        SELECT rid, date_of_note, note
        FROM progress_note_table
        WHERE trid IN ( ... patient_id_query ... )
        ORDER BY trid
    
        SELECT rid, date_of_diagnosis, diagnosis, diagnosis_description
        FROM diagnosis_table
        WHERE trid IN ( ... patient_id_query ... )
        ORDER BY trid

      This means we will repeat the patient_id_query, which may be inefficient.
      Options:
    
      - store the TRIDs in Python, then pass them as arguments
    
        - at which point the SQL string/packet length becomes relevant;
        - http://stackoverflow.com/questions/1869753/maximum-size-for-a-sql-server-query-in-clause-is-there-a-better-approach
        - http://stackoverflow.com/questions/16335011/what-is-maximum-query-size-for-mysql
        - http://stackoverflow.com/questions/96553/practical-limit-to-length-of-sql-query-specifically-mysql
      
      - let the database worry about it
    
        - probably best for now!

    - Display

      - One patient per page, with multiple results tables.

    - Boolean logic on patient selection
    
      - ... within
    
    """  # noqa
    def __init__(self,
                 output_columns: List[ColumnId] = None,
                 patient_conditions: List[WhereCondition] = None,
                 manual_patient_id_query: str = ''):
        """
        Args:
            output_columns:
                database columns that will be in the output, as
                list of :class:`crate_anon.common.sql.ColumnId` objects
            patient_conditions:
                restrictions on the patient, as a list of
                :class:`crate_anon.common.sql.WhereCondition` objects; they
                will be joined with ``AND``
            manual_patient_id_query:
                raw SQL; if specified, overrides ``patient_conditions`` and is
                used as the patient-finding part of the query; see
                :func:`set_override_query`
        """
        self._output_columns = output_columns or []  # type: List[ColumnId]
        self._patient_conditions = patient_conditions or []  # type: List[WhereCondition]  # noqa
        self._manual_patient_id_query = manual_patient_id_query or ''

    def __repr__(self) -> str:
        return (
            "<{qualname}("
            "output_columns={output_columns}, "
            "patient_conditions={patient_conditions}, "
            "manual_patient_id_query={manual_patient_id_query}"
            ") at {addr}>".format(
                qualname=self.__class__.__qualname__,
                output_columns=repr(self._output_columns),
                patient_conditions=repr(self._patient_conditions),
                manual_patient_id_query=repr(self._manual_patient_id_query),
                addr=hex(id(self)),
            )
        )

    def __eq__(self, other: 'PatientMultiQuery') -> bool:
        return (
            self._output_columns == other._output_columns and
            self._patient_conditions == other._patient_conditions and
            self._manual_patient_id_query == other._manual_patient_id_query
        )

    def __hash__(self) -> int:
        """
        WARNING: Python's hash() function converts the result of __hash__()
        to the integer width of the host machine, so 64-bit results can get
        down-converted to 32 bits. Use hash64() directly if you want a 64-bit
        result.
        """
        return self.hash64

    @property
    def hash64(self) -> int:
        """
        Return an integer (non-cryptographic) hash of the query.
        """
        return hash64(json_encode(self))

    @property
    def output_columns(self) -> List[ColumnId]:
        """
        Returns the output columns, as a list of
        :class:`crate_anon.common.sql.ColumnId` objects.
        """
        return self._output_columns

    @property
    def has_output_columns(self) -> bool:
        """
        Does this multiquery have any output columns?
        """
        return bool(self._output_columns)

    @property
    def ok_to_run(self) -> bool:
        """
        Is this OK to run, i.e. does it have a patient ID query and some output
        columns?
        """
        return self.has_output_columns and self.has_patient_id_query

    @property
    def patient_conditions(self) -> List[WhereCondition]:
        """
        Returns all ``WHERE`` conditions restricting the patient, as a list of
        :class:`crate_anon.common.sql.WhereCondition` objects.
        """
        return self._patient_conditions

    @property
    def manual_patient_id_query(self) -> str:
        """
        Returns the manual override SQL for the patient ID query.
        """
        return self._manual_patient_id_query

    def add_output_column(self, column_id: ColumnId) -> None:
        """
        Adds a database column to the output.
        """
        if column_id not in self._output_columns:
            self._output_columns.append(column_id)
            self._output_columns.sort()

    def clear_output_columns(self) -> None:
        """
        Removes all output columns from the multiquery.
        """
        self._output_columns = []

    def add_patient_condition(self, where: WhereCondition) -> None:
        """
        Adds a patient ``WHERE`` condition.

        Args:
            where: a :class:`crate_anon.common.sql.WhereCondition`
        """
        if where not in self._patient_conditions:
            self._patient_conditions.append(where)
            self._patient_conditions.sort()

    def clear_patient_conditions(self) -> None:
        """
        Removes all ``WHERE`` conditions on the patient.
        """
        self._patient_conditions = []

    def set_override_query(self, query: str) -> None:
        """
        Sets the manual override SQL for the patient ID query.

        Args:
            query: raw SQL

        This query should return a single column of MRID values that is fetched
        into Python and used to restrict other queries. Here's a fictional
        example to fetch the MRIDs for all patients who have the word
        "neutrophils" in their notes:

        .. code-block:: sql

            SELECT DISTINCT anonymous_output.patient.nhshash AS _mrid
            FROM anonymous_output.patient
            INNER JOIN anonymous_output.note ON anonymous_output.note.trid = anonymous_output.patient.trid
            WHERE MATCH (anonymous_output.note.note) AGAINST ('neutrophils')
                AND anonymous_output.patient.nhshash IS NOT NULL
            ORDER BY _mrid
            
        """  # noqa
        self._manual_patient_id_query = query

    def _get_select_mrid_column(self) -> Optional[ColumnId]:
        """
        Returns the MRID column from the first table in the patient ``WHERE``
        conditions, or ``None``.

        Returns:
            a :class:`crate_anon.common.sql.ColumnId` or ``None``

        """
        if not self._patient_conditions:
            return None
        return research_database_info.get_linked_mrid_column(
            self._patient_conditions[0].table_id)

    @property
    def has_patient_id_query(self) -> bool:
        """
        Does this multiquery have a patient ID query? This can either be one
        that the user has specified manually, or one built from ``WHERE``
        conditions that appears to refer to an MRID.
        """
        if self._manual_patient_id_query:
            return True
        if self._patient_conditions:
            mrid_col = self._get_select_mrid_column()
            if mrid_col and mrid_col.is_valid:
                return True
        return False

    def patient_id_query(self, with_order_by: bool = True) -> str:
        """
        Returns an SQL ``SELECT`` statement based on the list of ``WHERE``
        conditions already stored, joined with ``AND`` by default. (If a manual
        patient ID query has been specified, return that instead.)

        Args:
            with_order_by: add an ``ORDER BY`` query on the MRID; such an
                ordering is important for consistency across runs (but is
                prohibited by SQL Server in subqueries -- "The ORDER BY clause
                is invalid in views, inline functions, derived tables,
                subqueries, ... unless TOP, OFFSET or FOR XML is specified.")

        Returns:
            str: SQL

        """

        if self._manual_patient_id_query:
            # User has specified one manually.
            return self._manual_patient_id_query

        if not self._patient_conditions:
            return ''

        grammar = research_database_info.grammar
        select_mrid_column = self._get_select_mrid_column()
        if not select_mrid_column.is_valid:
            log.warning(
                f"PatientMultiQuery.patient_id_query(): invalid"
                f" select_mrid_column: {select_mrid_column!r}")
            # One way this can happen: (1) a user saves a PMQ; (2) the
            # administrator removes one of the databases!
            return ''
        mrid_alias = "_mrid"
        sql = add_to_select(
            '',
            grammar=grammar,
            select_elements=[SelectElement(column_id=select_mrid_column,
                                           alias=mrid_alias)],
            distinct=True,
            where_conditions=(self._patient_conditions + [WhereCondition(
                column_id=select_mrid_column, op="IS NOT NULL")]),
            where_type="AND",
            magic_join=True,
            formatted=False
        )
        if with_order_by:
            sql += " ORDER BY " + mrid_alias
        sql = format_sql(sql)
        # log.critical(sql)
        return sql

    @property
    def all_full_queries(self) -> List[TableQueryArgs]:
        """
        Returns all final queries. This is a list of multiple SQL queries, each
        retrieving information from one table, and all retrieving information
        for the same patient(s).

        The patients we use are defined by our :meth:`patient_id_query`.

        Returns:
            list: a list of :class:`TableQueryArgs` objects (q.v.)

        """
        return self.all_queries(mrids=None)

    def all_queries_specific_patients(
            self,
            mrids: List[int]) -> List[TableQueryArgs]:
        """
        Returns all final queries. This is a list of multiple SQL queries, each
        retrieving information from one table, and all retrieving information
        for the same patient(s).

        The patients we use are defined by the MRID list given.

        Args:
            mrids: list of MRIDs

        Returns:
            list: a list of :class:`TableQueryArgs` objects (q.v.)

        """
        return self.all_queries(mrids=mrids)

    def all_queries(self,
                    mrids: List[Any] = None) -> List[TableQueryArgs]:
        """
        Returns all final queries. This is a list of multiple SQL queries, each
        retrieving information from one table, and all retrieving information
        for the same patient(s).

        The patients we use are defined either by the MRID list given, or if
        that is empty or blank, our :meth:`patient_id_query`.

        Args:
            mrids: list of MRIDs; if this is ``None`` or empty, use the
                patients fetched (live) by our :meth:`patient_id_query`.

        Returns:
            list: a list of :class:`TableQueryArgs` objects (q.v.)

        """
        queries = []
        table_columns_map = columns_to_table_column_hierarchy(
            self._output_columns, sort=True)
        for table, columns in table_columns_map:
            table_sql_args = self.make_query(table_id=table,
                                             columns=columns,
                                             mrids=mrids)
            queries.append(table_sql_args)
        return queries

    def where_patient_clause(self, table_id: TableId,
                             grammar: SqlGrammar,
                             mrids: List[Any] = None) -> SqlArgsTupleType:
        """
        Returns an SQL WHERE clauses similar to ``sometable.mrid IN (1, 2, 3)``
        or ``sometable.mridcol IN (SELECT mrid FROM masterpatienttable)``. The
        clause is used to restrict patients by MRID.

        Args:
            table_id: :class:`crate_anon.common.sql.TableId` for the table
                whose MRID column we will apply the ``WHERE`` clause to
            grammar: :class:`cardinal_pythonlib.sql.sql_grammar.SqlGrammar`
                to use
            mrids: list of MRIDs; if this is ``None`` or empty, use the
                patients fetched (live) by our :meth:`patient_id_query`.

        Returns:
            tuple: ``sql, args``
        """
        mrid_column = research_database_info.get_mrid_column_from_table(
            table_id)
        if mrids:
            in_clause = ",".join(["?"] * len(mrids))
            # ... see notes for translate_sql_qmark_to_percent()
            args = mrids
        else:
            # If we haven't specified specific patients, use our patient-
            # finding query.
            in_clause = self.patient_id_query(with_order_by=False)
            # ... SQL Server moans if you use use ORDER BY in a subquery:
            # "The ORDER BY clause is invalid in views, inline functions,
            # derived tables, subqueries, ... unless TOP, OFFSET or FOR XML
            # is specified."
            args = []
        sql = f"{mrid_column.identifier(grammar)} IN ({in_clause})"
        return sql, args

    def make_query(self,
                   table_id: TableId,
                   columns: List[ColumnId],
                   mrids: List[Any] = None) -> TableQueryArgs:
        """
        Returns an SQL query to retrieve information from a single table for
        certain patients. This query is similar to ``SELECT a, b, c FROM
        sometable WHERE sometable.mrid IN (1, 2, 3)`` or ``SELECT a, b, c FROM
        sometable WHERE sometable.mrid IN (SELECT mrid FROM
        masterpatienttable)``. This then forms one query from (potentially)
        many for our patient(s).

        Args:
            table_id: a :class:`crate_anon.common.sql.TableId` for the SELECT
                FROM table
            columns: columns, specified as a list of
                :class:`crate_anon.common.sql.ColumnId`, to select from the
                table (in addition to which, we will always select the MRID
                column from that table)
            mrids: list of MRIDs; if this is ``None`` or empty, use the
                patients fetched (live) by our :meth:`patient_id_query`.

        Returns:
            a :class:`TableQueryArgs` object (q.v.)

        """
        if not columns:
            raise ValueError("No columns specified")
        grammar = research_database_info.grammar
        mrid_column = research_database_info.get_mrid_column_from_table(
            table_id)
        all_columns = [mrid_column]
        for c in columns:
            if c not in all_columns:
                all_columns.append(c)
        where_clause, args = self.where_patient_clause(table_id, grammar,
                                                       mrids)
        select_elements = [SelectElement(column_id=col)
                           for col in all_columns]
        where_conditions = [WhereCondition(raw_sql=where_clause)]
        sql = add_to_select('',
                            grammar=grammar,
                            select_elements=select_elements,
                            where_conditions=where_conditions,
                            magic_join=True,
                            formatted=True)
        return TableQueryArgs(table_id, sql, args)

    # -------------------------------------------------------------------------
    # Display
    # -------------------------------------------------------------------------

    @property
    def output_cols_html(self) -> str:
        """
        Returns all our output columns in HTML format.
        """
        grammar = research_database_info.grammar
        return prettify_sql_html("\n".join(
            [column_id.identifier(grammar)
             for column_id in self.output_columns]))

    @property
    def pt_conditions_html(self) -> str:
        """
        Returns all our patient WHERE conditions in HTML format.
        """
        grammar = research_database_info.grammar
        return prettify_sql_html("\nAND ".join([
            wc.sql(grammar) for wc in self.patient_conditions]))

    def summary_html(self, element_counter: HtmlElementCounter) -> str:
        """
        Returns an HTML representation of this multiquery.
        
        Args:
            element_counter: a
                :class:`crate_anon.crateweb.research.html_functions.HtmlElementCounter`,
                which will be modified

        Returns:
            str: HTML

        """  # noqa

        def collapser(x: str) -> str:
            return element_counter.overflow_div(contents=x)

        outcols = self.output_cols_html
        manual_query = self.manual_patient_id_query
        if manual_query:
            manual_or_auto = " (MANUAL)"
            ptselect = prettify_sql_html(manual_query)
        else:
            manual_or_auto = ""
            ptselect = self.pt_conditions_html
        return f"""
            Output columns:<br>
            {collapser(outcols)}
            Patient selection{manual_or_auto}:<br>
            {collapser(ptselect)}
        """

    # -------------------------------------------------------------------------
    # Data finder: COUNT(*) for all patient tables
    # -------------------------------------------------------------------------

    def gen_data_finder_queries(self, mrids: List[Any] = None) \
            -> Generator[TableQueryArgs, None, None]:
        """
        Generates a set of queries that, when executed, return the following
        summary columns from each of our tables, filtered for patients by our
        :meth:`where_patient_clause`, and grouped by ``master_research_id``
        (MRID):

        .. code-block:: sql

            master_research_id,
            table_name,
            COUNT(*) AS n_records,
            MIN(date_column) AS min_date,  -- NULL if no date column
            MAX(date_column) AS max_date   -- NULL if no date column

        These queries can be used to see quickly which tables have interesting
        information in.

        Args:
            mrids: list of MRIDs; if this is ``None`` or empty, use the
                patients fetched (live) by our :meth:`patient_id_query`.

        Yields:
            :class:`TableQueryArgs` objects (q.v.)

        """
        grammar = research_database_info.grammar
        mrid_alias = 'master_research_id'
        table_name_alias = 'table_name'
        n_records_alias = 'n_records'
        min_date_alias = 'min_date'
        max_date_alias = 'max_date'
        for table_id in research_database_info.get_mrid_linkable_patient_tables():  # noqa
            mrid_col = research_database_info.get_mrid_column_from_table(
                table=table_id)
            date_col = research_database_info.get_default_date_column(
                table=table_id)
            if date_col:
                min_date = f"MIN({date_col.identifier(grammar)})"
                max_date = f"MAX({date_col.identifier(grammar)})"
            else:
                min_date = "NULL"
                max_date = "NULL"
                # ... OK (at least in MySQL) to do:
                # SELECT col1, COUNT(*), NULL FROM table GROUP BY col1;
            where_clause, args = self.where_patient_clause(
                table_id, grammar, mrids)
            table_identifier = table_id.identifier(grammar)
            select_elements = [
                SelectElement(column_id=mrid_col, alias=mrid_alias),
                SelectElement(raw_select=sql_string_literal(table_identifier),
                              alias=table_name_alias),
                SelectElement(raw_select='COUNT(*)',
                              from_table_for_raw_select=table_id,
                              alias=n_records_alias),
                SelectElement(raw_select=min_date,
                              from_table_for_raw_select=table_id,
                              alias=min_date_alias),
                SelectElement(raw_select=max_date,
                              from_table_for_raw_select=table_id,
                              alias=max_date_alias),
            ]
            where_conditions = [WhereCondition(raw_sql=where_clause)]
            sql = add_to_select('',
                                grammar=grammar,
                                select_elements=select_elements,
                                where_conditions=where_conditions,
                                magic_join=True,
                                formatted=False)
            sql += "\nGROUP BY " + mrid_col.identifier(grammar)
            sql += "\nORDER BY " + mrid_alias
            sql = format_sql(sql)
            yield TableQueryArgs(table_identifier, sql, args)

    # -------------------------------------------------------------------------
    # Monster data: SELECT * for all patient tables
    # -------------------------------------------------------------------------

    def gen_monster_queries(self, mrids: List[int] = None) \
            -> Generator[TableQueryArgs, None, None]:
        """
        Generates a set of queries that, when executed, return ``SELECT *``
        from each of our tables, filtered for patients by our
        :meth:`where_patient_clause`. So it's like the basic Patient Explorer
        but with all columns in the output.

        These queries are used in the Patient Explorer "Monster Data" view.

        Args:
            mrids: list of MRIDs; if this is ``None`` or empty, use the
                patients fetched (live) by our :meth:`patient_id_query`.

        Yields:
            :class:`TableQueryArgs` objects (q.v.)

        """
        grammar = research_database_info.grammar
        for table_id in research_database_info.get_mrid_linkable_patient_tables():  # noqa
            mrid_col = research_database_info.get_mrid_column_from_table(
                table=table_id)
            where_clause, args = self.where_patient_clause(
                table_id, grammar, mrids)
            # We add the WHERE using our magic query machine, to get the joins
            # right:
            select_elements = [
                SelectElement(raw_select='*',
                              from_table_for_raw_select=table_id),
            ]
            where_conditions = [
                WhereCondition(raw_sql=where_clause,
                               from_table_for_raw_sql=mrid_col.table_id),
            ]
            sql = add_to_select(
                '',
                grammar=grammar,
                select_elements=select_elements,
                where_conditions=where_conditions,
                magic_join=True,
                formatted=False)
            sql += " ORDER BY " + mrid_col.identifier(grammar)
            sql = format_sql(sql)
            yield TableQueryArgs(table_id, sql, args)


# =============================================================================
# PatientExplorer
# =============================================================================

PATIENT_EXPLORER_FWD_REF = "PatientExplorer"


class PatientExplorer(models.Model):
    """
    Class to explore the research database on a per-patient basis.
    """
    class Meta:
        app_label = "research"

    id = models.AutoField(primary_key=True)  # automatic
    user = models.ForeignKey(settings.AUTH_USER_MODEL,
                             on_delete=models.CASCADE)
    patient_multiquery = JsonClassField(
        verbose_name='PatientMultiQuery as JSON',
        null=True)  # type: PatientMultiQuery
    pmq_hash = models.BigIntegerField(
        verbose_name='64-bit non-cryptographic hash of JSON of '
                     'patient_multiquery')
    active = models.BooleanField(default=True)  # see save() below
    created = models.DateTimeField(auto_now_add=True)
    deleted = models.BooleanField(
        default=False,
        verbose_name="Deleted from the user's perspective. "
                     "Audited queries are never properly deleted.")
    audited = models.BooleanField(default=False)

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        if not self.patient_multiquery:
            self.patient_multiquery = PatientMultiQuery()

    def __str__(self) -> str:
        return f"<PatientExplorer id={self.id}>"

    def save(self, *args, **kwargs) -> None:
        """
        Custom save method. Ensures that only one :class:`PatientExplorer` has
        ``active == True`` for a given user. Also sets the hash.
        """
        if self.active:
            PatientExplorer.objects\
                .filter(user=self.user, active=True)\
                .update(active=False)
        self.pmq_hash = self.patient_multiquery.hash64
        # Beware: Python's hash() function will downconvert to 32 bits on 32-bit
        # machines; use pmq.hash64() directly, not hash(pmq).
        super().save(*args, **kwargs)

    # -------------------------------------------------------------------------
    # Fetching
    # -------------------------------------------------------------------------

    @staticmethod
    def get_active_pe_or_none(request: HttpRequest) \
            -> Optional[PATIENT_EXPLORER_FWD_REF]:
        """
        Args:
            request: the :class:`django.http.request.HttpRequest`

        Returns:
            The active :class:`PatientExplorer` for the user, or ``None``.

        """
        if not request.user.is_authenticated:
            return None
        try:
            return PatientExplorer.objects.get(user=request.user, active=True)
        except PatientExplorer.DoesNotExist:
            return None

    @staticmethod
    def get_active_pe_id_or_none(request: HttpRequest) -> Optional[int]:
        """
        Args:
            request: the :class:`django.http.request.HttpRequest`

        Returns:
            The integer PK of the active :class:`PatientExplorer` for the user,
            or ``None``.

        """
        if not request.user.is_authenticated:
            return None
        try:
            pe = PatientExplorer.objects.get(user=request.user, active=True)
            return pe.id
        except PatientExplorer.DoesNotExist:
            return None

    # -------------------------------------------------------------------------
    # Activating, deleting, auditing
    # -------------------------------------------------------------------------

    def activate(self) -> None:
        """
        Activates this :class:`PatientExplorer` (and deactivates any others).
        """
        self.active = True
        self.save()

    def mark_audited(self) -> None:
        """
        Mark the query as having been executed and audited. (This prevents it
        from being wholly deleted.)
        """
        if self.audited:
            return
        self.audited = True
        self.save()

    def mark_deleted(self) -> None:
        """
        Mark the query as deleted.

        This will stop it being shown. It will not delete it from the database.

        We use this deletion method for queries that have been executed, so
        need an audit trail.
        """
        if self.deleted:
            # log.debug("pointless")
            return
        self.deleted = True
        self.active = False
        # log.debug("about to save")
        self.save()
        # log.debug("saved")

    def delete_if_permitted(self) -> None:
        """
        Delete the query.

        - If a query has been executed and therefore audited, it isn't properly
          deleted; it's just marked as deleted.
        - If a query has never been executed, we can delete it entirely.
        """
        if self.deleted:
            log.debug("already flagged as deleted")
            return
        if self.audited:
            log.debug("marking as deleted")
            self.mark_deleted()
        else:
            # actually delete
            log.debug("actually deleting")
            self.delete()

    def audit(self, count_only: bool = False, n_records: int = 0,
              failed: bool = False, fail_msg: str = "") -> None:
        """
        Audit the execution of this query:

        - insert an audit entry referring to this query
        - mark the query as having been audited (so it's not deleted)

        Args:
            count_only: did we know (in advance) that this was a
                ``COUNT()``-only query?
            n_records: how many records were returned?
            failed: did the query fail?
            fail_msg: if the query failed, the associated failure message
        """
        a = PatientExplorerAudit(patient_explorer=self,
                                 count_only=count_only,
                                 n_records=n_records,
                                 failed=failed,
                                 fail_msg=fail_msg)
        a.save()
        self.mark_audited()

    # -------------------------------------------------------------------------
    # Using the internal PatientMultiQuery
    # -------------------------------------------------------------------------

    def all_queries(self,
                    mrids: List[Any] = None) -> List[TableQueryArgs]:
        """
        Returns all queries from our :attr:`patient_multiquery`. See
        :meth:`PatientMultiQuery.all_queries`

        Args:
            mrids: list of MRIDs; if this is ``None`` or empty, use the
                patients fetched (live) by our :attr:`patient_multiquery`'s
                :meth:`PatientMultiQuery.patient_id_query`.

        Returns:
            list: a list of :class:`TableQueryArgs` objects (q.v.)

        """
        return self.patient_multiquery.all_queries(mrids=mrids)

    @staticmethod
    def get_executed_cursor(sql: str, args: List[Any] = None) -> CursorWrapper:
        """
        Executes a query (via the research database) and returns its cursor.

        Args:
            sql: SQL text
            args: arguments to SQL query

        Returns:
            a :class:`django.db.backends.utils.CursorWrapper`, which is a
            context manager that behaves as the executed cursor and also closes
            it on completion
        """
        sql = translate_sql_qmark_to_percent(sql)
        cursor = get_executed_researchdb_cursor(sql, args)
        return cursor

    def get_patient_mrids(self) -> List[int]:
        """
        Returns all MRIDs from our :attr:`patient_multiquery`'s
        :meth:`PatientMultiQuery.patient_id_query`.
        """
        sql = self.patient_multiquery.patient_id_query(with_order_by=True)
        # log.critical(sql)
        with self.get_executed_cursor(sql) as cursor:
            return [row[0] for row in cursor.fetchall()]

    def get_zipped_tsv_binary(self) -> bytes:
        """
        Returns a ZIP file containing TSVs, one for each table in our
        :attr:`patient_multiquery`.
        """
        # Don't pass giant result sets around beyond what's necessary.
        # Use cursor.fetchone()
        grammar = make_grammar(settings.RESEARCH_DB_DIALECT)
        memfile = io.BytesIO()
        z = zipfile.ZipFile(memfile, "w")
        for tsa in self.patient_multiquery.all_queries():
            table_id = tsa.table_id
            sql = tsa.sql
            args = tsa.args
            with self.get_executed_cursor(sql, args) as cursor:
                try:
                    fieldnames = get_fieldnames_from_cursor(cursor)
                except TypeError:
                    fieldnames = []
                tsv = make_tsv_row(fieldnames)
                row = cursor.fetchone()
                while row is not None:
                    tsv += make_tsv_row(row)
                    row = cursor.fetchone()
            filename = table_id.identifier(grammar) + ".tsv"
            z.writestr(filename, tsv.encode("utf-8"))
        z.close()
        return memfile.getvalue()

    def get_xlsx_binary(self) -> bytes:
        """
        Returns an XLSX (Excel) file containing spreadsheets, one for each
        table in our :attr:`patient_multiquery`.

        Other notes:

        - cell size: see
          http://stackoverflow.com/questions/13197574/python-openpyxl-column-width-size-adjust;
          and the "auto_size" / "bestFit" options don't really do the job,
          according to the interweb.

        """  # noqa
        wb = Workbook()
        wb.remove_sheet(wb.active)  # remove the autocreated blank sheet
        sqlsheet_rows = [["Table", "SQL", "Args", "Executed_at"]]
        for tsa in self.patient_multiquery.all_queries():
            table_id = tsa.table_id
            sql = tsa.sql
            args = tsa.args
            sqlsheet_rows.append([str(table_id), sql, repr(args),
                                  datetime.datetime.now()])
            ws = wb.create_sheet(title=str(table_id))
            with self.get_executed_cursor(sql, args) as cursor:
                try:
                    fieldnames = get_fieldnames_from_cursor(cursor)
                except (AttributeError, IndexError):
                    fieldnames = []
                ws.append(fieldnames)
                row = cursor.fetchone()
                while row is not None:
                    ws.append(gen_excel_row_elements(ws, row))
                    row = cursor.fetchone()
        sql_ws = wb.create_sheet(title="SQL")
        for r in sqlsheet_rows:
            sql_ws.append(r)
        return excel_to_bytes(wb)

    # -------------------------------------------------------------------------
    # Using the internal PatientMultiQuery
    # -------------------------------------------------------------------------

    def get_patient_id_query(self, with_order_by: bool = True) -> str:
        """
        Returns SQL from our :attr:`patient_multiquery`'s
        :meth:`PatientMultiQuery.patient_id_query` (q.v.).

        Args:
            with_order_by: see :meth:`PatientMultiQuery.patient_id_query`
        """
        return self.patient_multiquery.patient_id_query(
            with_order_by=with_order_by)

    # -------------------------------------------------------------------------
    # Display
    # -------------------------------------------------------------------------

    @property
    def summary_html(self) -> str:
        """
        Return HTML summarizing this object.
        """
        # Nasty hack. We want collapsing things, so we want HTML element IDs.
        # We could build the HTML table in code for the Patient Explorer
        # chooser, but I was trying to do it in Django templates.
        # However, it's not easy to pass parameters (such as an
        # HtmlElementCounter) back to Python from Django templates.
        # So we can hack it a bit:
        element_counter = HtmlElementCounter(prefix=f"pe_{self.id}_")
        return self.patient_multiquery.summary_html(
            element_counter=element_counter)

    @property
    def has_patient_id_query(self) -> bool:
        """
        Does our our :attr:`patient_multiquery` have a patient ID query?

        See :meth:`PatientMultiQuery.has_patient_id_query`.
        """
        return self.patient_multiquery.has_patient_id_query

    @property
    def has_output_columns(self) -> bool:
        """
        Does our our :attr:`patient_multiquery` have output columns?

        See :meth:`PatientMultiQuery.has_output_columns`.
        """
        return self.patient_multiquery.has_output_columns

    # -------------------------------------------------------------------------
    # Data finder
    # -------------------------------------------------------------------------

    @property
    def data_finder_excel(self) -> bytes:
        """
        Returns an XSLX (Excel) file containing summary (count) information
        for each table.

        See :meth:`PatientMultiQuery.gen_data_finder_queries`.
        """
        fieldnames = []
        wb = Workbook()
        wb.remove_sheet(wb.active)  # remove the autocreated blank sheet
        all_ws = wb.create_sheet("All_patients")
        sql_ws = wb.create_sheet("SQL")
        sql_ws.append(["Table", "SQL", "Args", "Executed_at"])

        for tsa in self.patient_multiquery.gen_data_finder_queries():
            table_identifier = tsa.table_id
            sql = tsa.sql
            args = tsa.args
            sql_ws.append([table_identifier,
                           format_sql(sql),
                           repr(args),
                           datetime.datetime.now()])
            with self.get_executed_cursor(sql, args) as cursor:
                if not fieldnames:
                    try:
                        fieldnames = get_fieldnames_from_cursor(cursor)
                    except TypeError:
                        fieldnames = []
                    all_ws.append(fieldnames)
                row = cursor.fetchone()
                while row is not None:
                    mrid = str(row[0])
                    if mrid in wb:
                        ws = wb[mrid]
                    else:
                        ws = wb.create_sheet(mrid)
                        ws.append(fieldnames)
                    rowtuple = tuple(row)
                    ws.append(rowtuple)
                    all_ws.append(rowtuple)
                    row = cursor.fetchone()
        return excel_to_bytes(wb)


# =============================================================================
# PatientExplorer auditing class
# =============================================================================

class PatientExplorerAudit(models.Model):
    """
    Audit log for a PatientExplorer.
    """
    id = models.AutoField(primary_key=True)  # automatic
    patient_explorer = models.ForeignKey('PatientExplorer',
                                         on_delete=models.PROTECT)
    when = models.DateTimeField(auto_now_add=True)
    count_only = models.BooleanField(default=False)
    n_records = models.IntegerField(default=0)
    # ... not PositiveIntegerField; SQL Server gives -1, for example
    failed = models.BooleanField(default=False)
    fail_msg = models.TextField()

    def __str__(self):
        return f"<PatientExplorerAudit id={self.id}>"
